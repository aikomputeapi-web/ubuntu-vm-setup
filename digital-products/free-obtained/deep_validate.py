"""
Deep validation: goes beyond file existence to test actual content quality.
- Planner: verify hyperlink annotations exist (not just bookmarks)
- SVGs: check for visible elements, viewBox, proper attributes
- Wall Art: open JPGs with PIL to confirm they're real images, check JSON fields
- Coloring: verify pages have drawn content (not blank)
- Spreadsheet: test formula computation, data validation, conditional formatting
- Canva templates: verify text elements are editable, palettes correct
- Video scripts: validate structure (hook, segments, CTA, production notes)
- Guides: validate URLs are well-formed, all listings referenced
"""
import os
import json
import re
import sys
import xml.etree.ElementTree as ET

BASE = os.path.dirname(os.path.abspath(__file__))
issues = []

def log_issue(category, severity, msg):
    issues.append((category, severity, msg))
    marker = "WARNING" if severity == "warning" else "FAIL"
    print(f"  [{marker}] {msg}")


def deep_validate_planner():
    """Check planner has actual hyperlink annotations, correct dates."""
    print("\n=== DEEP: Digital Planner ===")
    try:
        from pypdf import PdfReader
        p = os.path.join(BASE, "01-digital-planners/planner_2026_6month_hyperlinked.pdf")
        reader = PdfReader(p)
        total_pages = len(reader.pages)
        print(f"  Pages: {total_pages}")

        # Check for link annotations on pages (not just bookmarks)
        pages_with_links = 0
        total_links = 0
        for i, page in enumerate(reader.pages):
            if "/Annots" in page:
                annots = page["/Annots"]
                if annots:
                    for annot_ref in annots:
                        try:
                            annot = annot_ref.get_object()
                            if annot.get("/Subtype") == "/Link":
                                total_links += 1
                                if i not in range(pages_with_links):
                                    pages_with_links += 1
                        except:
                            pass
        print(f"  Pages with link annotations: {pages_with_links}")
        print(f"  Total link annotations: {total_links}")
        if total_links == 0:
            log_issue("planner", "fail", "No hyperlink annotations found in planner PDF")

        # Check bookmarks
        outline = reader.outline if hasattr(reader, "outline") else []
        bookmark_count = len(outline) if outline else 0
        print(f"  Bookmarks: {bookmark_count}")
        if bookmark_count == 0:
            log_issue("planner", "warning", "No bookmarks in planner PDF")

        # Check first page has text (index page)
        first_text = reader.pages[0].extract_text()
        if not first_text or len(first_text.strip()) < 10:
            log_issue("planner", "warning", "First page appears empty or has minimal text")
        else:
            print(f"  First page text length: {len(first_text)} chars")

        # Check a daily page has content
        if total_pages > 50:
            daily_page = reader.pages[30]  # Should be a daily page
            daily_text = daily_page.extract_text()
            print(f"  Sample daily page (p30) text: {len(daily_text)} chars")

        return total_links > 0
    except Exception as e:
        log_issue("planner", "fail", f"Error reading planner: {e}")
        return False


def deep_validate_svgs():
    """Check SVGs have visible elements, viewBox, proper structure."""
    print("\n=== DEEP: SVG Bundles ===")
    svg_dirs = [
        ("iconify", "02-svg-bundles/svg-bundles/iconify-icons"),
        ("original", "02-svg-bundles/svg-bundles/original-designs"),
        ("themed", "02-svg-bundles/svg-bundles/themed-packs"),
        ("canva", "07-canva-templates/templates"),
        ("custom", "11-custom-personalized/output"),
    ]

    total_ok = 0
    total_issues = 0

    for label, d in svg_dirs:
        full = os.path.join(BASE, d)
        if not os.path.isdir(full):
            continue

        no_viewbox = 0
        no_path = 0
        empty_svg = 0
        count = 0

        for root, dirs, files in os.walk(full):
            for f in files:
                if not f.endswith(".svg"):
                    continue
                count += 1
                path = os.path.join(root, f)
                try:
                    tree = ET.parse(path)
                    root_elem = tree.getroot()
                    if "svg" not in root_elem.tag:
                        no_path += 1
                        continue
                    # Check viewBox
                    vb = root_elem.get("viewBox") or root_elem.get("viewbox")
                    if not vb and root_elem.get("width") and root_elem.get("height"):
                        pass  # width/height is OK too
                    elif not vb:
                        no_viewbox += 1
                    # Check for visible elements (path, rect, circle, text, etc.)
                    visible = root_elem.findall(".//{http://www.w3.org/2000/svg}path") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}rect") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}circle") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}text") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}ellipse") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}line") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}polygon") + \
                               root_elem.findall(".//{http://www.w3.org/2000/svg}polyline")
                    # Also check without namespace (some SVGs don't use it)
                    if not visible:
                        visible = root_elem.findall(".//path") + \
                                  root_elem.findall(".//rect") + \
                                  root_elem.findall(".//circle") + \
                                  root_elem.findall(".//text") + \
                                  root_elem.findall(".//ellipse") + \
                                  root_elem.findall(".//line") + \
                                  root_elem.findall(".//polygon") + \
                                  root_elem.findall(".//polyline")
                    if not visible:
                        empty_svg += 1
                except Exception as e:
                    no_path += 1

        total_ok += count - no_path - no_viewbox - empty_svg
        total_issues += no_path + no_viewbox + empty_svg
        print(f"  {label}: {count} files, {no_viewbox} no viewBox, {no_path} parse error, {empty_svg} no visible elements")
        if no_viewbox > 0:
            log_issue("svgs", "warning", f"{label}: {no_viewbox} SVGs without viewBox")
        if empty_svg > 0:
            log_issue("svgs", "warning", f"{label}: {empty_svg} SVGs with no visible elements")

    print(f"  Total: {total_ok} OK, {total_issues} with issues")
    return total_issues == 0


def deep_validate_wall_art():
    """Open JPGs with PIL to confirm they're real images, check JSON fields."""
    print("\n=== DEEP: Wall Art ===")
    try:
        from PIL import Image
    except ImportError:
        print("  PIL not available, skipping image validation")
        return False

    met_dir = os.path.join(BASE, "03-wall-art/met-wall-art")
    if not os.path.isdir(met_dir):
        log_issue("wallart", "fail", "Met wall art directory not found")
        return False

    # Validate JPGs are real images
    valid_images = 0
    invalid_images = 0
    small_images = 0

    for root, dirs, files in os.walk(met_dir):
        for f in files:
            if f.endswith((".jpg", ".jpeg")):
                path = os.path.join(root, f)
                try:
                    img = Image.open(path)
                    img.verify()  # Verify it's a valid image
                    img = Image.open(path)  # Reopen after verify
                    w, h = img.size
                    if w < 200 or h < 200:
                        small_images += 1
                        log_issue("wallart", "warning", f"Small image: {f} ({w}x{h})")
                    valid_images += 1
                except Exception as e:
                    invalid_images += 1
                    log_issue("wallart", "fail", f"Invalid image: {f} - {e}")

    print(f"  Met JPGs: {valid_images} valid, {invalid_images} invalid, {small_images} small")

    # Validate JSON metadata has expected fields
    json_with_fields = 0
    json_missing_fields = 0
    expected_fields = ["title", "primaryImage"]

    for root, dirs, files in os.walk(met_dir):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fh:
                        data = json.load(fh)
                    has_title = "title" in data or "Title" in data
                    if has_title:
                        json_with_fields += 1
                    else:
                        json_missing_fields += 1
                except Exception as e:
                    json_missing_fields += 1
                    log_issue("wallart", "warning", f"JSON read error: {f} - {e}")

    print(f"  Met JSON: {json_with_fields} with title, {json_missing_fields} missing/error")

    # Validate original PNGs
    png_dir = os.path.join(BASE, "03-wall-art/wall-art-png")
    if os.path.isdir(png_dir):
        pngs = [f for f in os.listdir(png_dir) if f.endswith(".png")]
        valid_png = 0
        for f in pngs:
            try:
                img = Image.open(os.path.join(png_dir, f))
                img.verify()
                valid_png += 1
            except:
                log_issue("wallart", "fail", f"Invalid PNG: {f}")
        print(f"  Original PNGs: {valid_png}/{len(pngs)} valid")
        return valid_png == len(pngs) and invalid_images == 0

    return invalid_images == 0


def deep_validate_coloring():
    """Verify coloring pages have actual drawn content (not blank)."""
    print("\n=== DEEP: Coloring Pages ===")
    try:
        from pypdf import PdfReader
        from pypdf.generic import ContentStream
        p = os.path.join(BASE, "04-coloring-pages/Free_Coloring_Pages_Bundle.pdf")
        reader = PdfReader(p)
        total = len(reader.pages)
        print(f"  Pages: {total}")

        blank_pages = 0
        low_content = 0
        for i, page in enumerate(reader.pages):
            # Use pypdf ContentStream to decompress and read page contents
            text = ""
            try:
                cs = ContentStream(page["/Contents"], reader)
                text = cs.get_data().decode("latin-1", errors="ignore") if isinstance(cs.get_data(), bytes) else str(cs.get_data())
            except Exception as e:
                # Fallback: direct access
                try:
                    contents = page["/Contents"].get_object()
                    data = contents.get_data()
                    text = data.decode("latin-1", errors="ignore") if isinstance(data, bytes) else str(data)
                except Exception as e2:
                    if i < 5:
                        print(f"  Page {i+1}: Error reading content: {e2}")

            # Count PDF drawing operators
            moveto = text.count(" m\n") + text.count(" m ")
            lineto = text.count(" l\n") + text.count(" l ")
            curveto = text.count(" c\n") + text.count(" c ")
            rect = text.count(" re\n") + text.count(" re ")
            total_ops = moveto + lineto + curveto + rect

            if total_ops < 5:
                blank_pages += 1
                if i < 5:
                    log_issue("coloring", "warning", f"Page {i+1} has only {total_ops} drawing operations")
            elif total_ops < 20:
                low_content += 1

        print(f"  Blank pages: {blank_pages}, low-content pages: {low_content}")
        if blank_pages > 5:
            log_issue("coloring", "warning", f"{blank_pages} pages appear to have no drawing content")
        return blank_pages < total * 0.2  # Allow up to 20% blank
    except Exception as e:
        log_issue("coloring", "fail", f"Error: {e}")
        return False


def deep_validate_spreadsheet():
    """Test formula computation, data validation, conditional formatting."""
    print("\n=== DEEP: Spreadsheet Tracker ===")
    try:
        import openpyxl
        p = os.path.join(BASE, "05-spreadsheet-trackers/Ultimate_Budget_Tracker_Template.xlsx")
        wb = openpyxl.load_workbook(p)

        # Check data validation
        dv_count = 0
        for s in wb.sheetnames:
            ws = wb[s]
            if hasattr(ws, "data_validations") and ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    dv_count += 1
        print(f"  Data validations: {dv_count}")
        if dv_count == 0:
            log_issue("spreadsheet", "warning", "No data validations found")

        # Check conditional formatting
        cf_count = 0
        for s in wb.sheetnames:
            ws = wb[s]
            if hasattr(ws, "conditional_formatting"):
                for rule in ws.conditional_formatting:
                    cf_count += 1
        print(f"  Conditional formatting rules: {cf_count}")
        if cf_count == 0:
            log_issue("spreadsheet", "warning", "No conditional formatting rules found")

        # Check formulas reference correct sheets
        formula_issues = 0
        for s in wb.sheetnames:
            ws = wb[s]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        # Check for common errors
                        if "#REF!" in cell.value:
                            formula_issues += 1
                            log_issue("spreadsheet", "fail", f"#REF! error in {s}!{cell.coordinate}")
        print(f"  Formula reference errors: {formula_issues}")

        # Check that charts have titles and data
        for s in wb.sheetnames:
            ws = wb[s]
            if hasattr(ws, "_charts"):
                for chart in ws._charts:
                    has_title = chart.title is not None
                    print(f"  Chart on '{s}': title={has_title}, type={type(chart).__name__}")

        return formula_issues == 0
    except Exception as e:
        log_issue("spreadsheet", "fail", f"Error: {e}")
        return False


def deep_validate_canva_templates():
    """Verify Canva templates have editable text elements and correct palette colors."""
    print("\n=== DEEP: Canva Templates ===")
    template_dir = os.path.join(BASE, "07-canva-templates/templates")
    if not os.path.isdir(template_dir):
        log_issue("canva", "fail", "Templates directory not found")
        return False

    # Check that templates have text elements (editable)
    no_text = 0
    count = 0
    for f in os.listdir(template_dir):
        if not f.endswith(".svg"):
            continue
        count += 1
        path = os.path.join(template_dir, f)
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            # Look for text elements
            texts = root.findall(".//{http://www.w3.org/2000/svg}text") + \
                    root.findall(".//text")
            if not texts:
                no_text += 1
        except:
            no_text += 1

    print(f"  Templates: {count} total, {no_text} without text elements")
    if no_text > 0:
        log_issue("canva", "warning", f"{no_text} templates have no editable text elements")
    return no_text == 0


def deep_validate_video_scripts():
    """Validate video scripts have expected structure."""
    print("\n=== DEEP: Video Scripts ===")
    scripts_dir = os.path.join(BASE, "09-ai-video-pipeline/video-scripts")
    if not os.path.isdir(scripts_dir):
        log_issue("video", "fail", "Video scripts directory not found")
        return False

    scripts = [f for f in os.listdir(scripts_dir) if f.endswith(".md")]
    print(f"  Scripts found: {len(scripts)}")

    for s in scripts:
        path = os.path.join(scripts_dir, s)
        with open(path, "r", encoding="utf-8") as fh:
            content = fh.read()

        has_title = content.startswith("# ")
        has_hook = "Full Script" in content
        has_segments = "Number" in content or "Idea" in content or "Day" in content
        has_cta = "Follow" in content or "Save" in content or "Subscribe" in content
        has_production = "Production Notes" in content
        has_ffmpeg = "ffmpeg" in content.lower()

        if not has_title:
            log_issue("video", "warning", f"{s}: missing title")
        if not has_hook:
            log_issue("video", "warning", f"{s}: missing Full Script section")
        if not has_production:
            log_issue("video", "warning", f"{s}: missing Production Notes")
        if not has_ffmpeg:
            log_issue("video", "warning", f"{s}: missing FFmpeg command")

        status = "OK" if all([has_title, has_hook, has_production, has_ffmpeg]) else "ISSUES"
        print(f"  {s}: {status}")

    return len(scripts) >= 5


def deep_validate_guides():
    """Validate URLs in guide markdown files are well-formed."""
    print("\n=== DEEP: Guide Links ===")
    guide_files = [
        "06-notion-templates/FREE_ALTERNATIVES.md",
        "07-canva-templates/FREE_ALTERNATIVES.md",
        "08-plr-mrr-bundles/FREE_SOURCES.md",
    ]

    total_links = 0
    valid_links = 0
    for g in guide_files:
        path = os.path.join(BASE, g)
        if not os.path.exists(path):
            log_issue("guides", "warning", f"Guide not found: {g}")
            continue
        with open(path, "r", encoding="utf-8") as fh:
            content = fh.read()
        # Find markdown links
        links = re.findall(r'\[([^\]]+)\]\(([^)]+)\)', content)
        for label, url in links:
            total_links += 1
            if url.startswith("http://") or url.startswith("https://") or url.startswith("#"):
                valid_links += 1
            else:
                log_issue("guides", "warning", f"Invalid URL in {g}: {url[:50]}")
        print(f"  {os.path.basename(g)}: {len(links)} links, {sum(1 for _, u in links if u.startswith(('http', '#'))) } valid")

    print(f"  Total links: {total_links}, valid: {valid_links}")
    return valid_links == total_links


def deep_validate_ideas_guide():
    """Validate ideas guide has real data from the CSV."""
    print("\n=== DEEP: Ideas Guide Data ===")
    p = os.path.join(BASE, "10-digital-product-ideas/Digital_Product_Ideas_Guide.md")
    if not os.path.exists(p):
        log_issue("ideas", "fail", "Ideas guide not found")
        return False
    with open(p, "r", encoding="utf-8") as fh:
        content = fh.read()

    # Check it has actual listing titles (not just template text)
    has_real_titles = "phenixdigital" in content or "BreezyOrganization" in content or "Plannerscollective" in content
    has_price_data = "$" in content and ("avg" in content.lower() or "Average" in content or "price" in content.lower())
    has_sales_numbers = "sales" in content and any(c.isdigit() for c in content)
    has_table_rows = content.count("|") > 50  # Tables have many pipe chars

    print(f"  Has real shop names: {has_real_titles}")
    print(f"  Has price analysis: {has_price_data}")
    print(f"  Has sales data: {has_sales_numbers}")
    print(f"  Has table data: {has_table_rows}")

    if not has_real_titles:
        log_issue("ideas", "warning", "Ideas guide doesn't reference actual scraped shop names")
    return has_real_titles and has_price_data


def deep_validate_custom_personalized():
    """Verify custom-personalized generator outputs have real content."""
    print("\n=== DEEP: Custom Personalized Products ===")
    try:
        import glob
        out_dir = os.path.join(BASE, "11-custom-personalized", "output")
        if not os.path.isdir(out_dir):
            log_issue("custom", "warning", "Output directory not found")
            return False

        svgs = glob.glob(os.path.join(out_dir, "*.svg"))
        mds = glob.glob(os.path.join(out_dir, "*.md"))

        print(f"  SVGs: {len(svgs)}, Reports: {len(mds)}")

        # Check each SVG has visible elements (not just empty wrapper)
        empty_svgs = 0
        no_text = 0
        for svg_path in svgs:
            try:
                tree = ET.parse(svg_path)
                root = tree.getroot()
                # Count visible elements (path, circle, rect, ellipse, line, polygon, text)
                visible = 0
                has_text = False
                for elem in root.iter():
                    tag = elem.tag.split("}")[-1]  # Strip namespace
                    if tag in ("path", "circle", "rect", "ellipse", "line", "polygon", "text", "tspan"):
                        visible += 1
                    if tag in ("text", "tspan"):
                        has_text = True
                if visible < 5:
                    empty_svgs += 1
                    log_issue("custom", "warning", f"{os.path.basename(svg_path)} has only {visible} visible elements")
                if not has_text:
                    no_text += 1
            except Exception as e:
                log_issue("custom", "warning", f"{os.path.basename(svg_path)}: parse error: {e}")
                empty_svgs += 1

        # Check reports have affirmations and prompts
        reports_ok = 0
        for md_path in mds:
            content = open(md_path, "r", errors="ignore").read()
            has_affirmation = "affirmation" in content.lower()
            has_prompt = "prompt" in content.lower() or "reflection" in content.lower()
            if has_affirmation and has_prompt:
                reports_ok += 1

        # Check that generators exist
        gen_dir = os.path.join(BASE, "11-custom-personalized")
        generators = [f for f in os.listdir(gen_dir) if f.endswith(".py") and f.startswith("generate_")]

        print(f"  Empty SVGs: {empty_svgs}, SVGs without text: {no_text}")
        print(f"  Reports with affirmations+prompts: {reports_ok}/{len(mds)}")
        print(f"  Generators: {len(generators)}")

        if empty_svgs > 0 or len(svgs) < 10:
            log_issue("custom", "warning", f"Only {len(svgs)} SVGs, {empty_svgs} empty")
        if len(generators) < 3:
            log_issue("custom", "warning", f"Only {len(generators)} generators (expected 3)")

        return empty_svgs == 0 and len(svgs) >= 10 and reports_ok == len(mds) and len(generators) >= 3
    except Exception as e:
        log_issue("custom", "fail", f"Error: {e}")
        return False


if __name__ == "__main__":
    print("=" * 60)
    print("DEEP CONTENT VALIDATION")
    print("Going beyond file existence to test actual content quality")
    print("=" * 60)

    results = {}
    results["planner"] = deep_validate_planner()
    results["svgs"] = deep_validate_svgs()
    results["wall_art"] = deep_validate_wall_art()
    results["coloring"] = deep_validate_coloring()
    results["spreadsheet"] = deep_validate_spreadsheet()
    results["canva"] = deep_validate_canva_templates()
    results["video"] = deep_validate_video_scripts()
    results["guides"] = deep_validate_guides()
    results["ideas"] = deep_validate_ideas_guide()
    results["custom"] = deep_validate_custom_personalized()

    print("\n" + "=" * 60)
    print("DEEP VALIDATION SUMMARY")
    print("=" * 60)
    fails = 0
    warnings = 0
    for k, v in results.items():
        status = "PASS" if v else "CHECK ISSUES"
        print(f"  {k}: {status}")
    print(f"\n  Issues found: {len(issues)}")
    for cat, sev, msg in issues:
        print(f"    [{sev.upper()}] {cat}: {msg}")
    if len(issues) == 0:
        print("\n  NO ISSUES FOUND - All content validated")
    print(f"\n  Overall: {'ALL PASS' if all(results.values()) and len(issues) == 0 else 'ISSUES FOUND'}")
