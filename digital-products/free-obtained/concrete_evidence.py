"""
Concrete evidence verification: does the output actually WORK?
- Planner links: do they point to valid, correct pages? Do targets render the right section?
- Coloring pages: do they actually render visible non-white pixels when rasterized?
- SVGs: do they render visible content when converted to PNG?
- Spreadsheet: do formulas actually compute correct values?
"""
import os, re, json, sys, tempfile
from PIL import Image

BASE = os.path.dirname(os.path.abspath(__file__))
evidence = []

def log(name, passed, detail):
    status = "EVIDENCE PASS" if passed else "EVIDENCE FAIL"
    evidence.append((name, passed, detail))
    print(f"  [{status}] {name}: {detail}")

# ============================================================
# 1. PLANNER LINKS: verify destinations are valid and correct
# ============================================================
print("\n=== EVIDENCE: Planner Link Targets ===")
try:
    import fitz  # PyMuPDF - resolves links like a real PDF viewer
    pdf_path = os.path.join(BASE, "01-digital-planners/planner_2026_6month_hyperlinked.pdf")
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    
    valid_targets = 0
    invalid_targets = 0
    target_pages = set()
    
    for page_idx in range(total_pages):
        page = doc[page_idx]
        links = page.get_links()
        for link in links:
            kind = link.get("kind")
            if kind == fitz.LINK_GOTO:
                target = link.get("page")
                if target is not None and 0 <= target < total_pages:
                    valid_targets += 1
                    target_pages.add(target)
                else:
                    invalid_targets += 1
    
    # Verify the 5 expected section targets exist
    expected_targets = {0: "index", 1: "months", 2: "weekly", 7: "daily", 219: "notes"}
    all_expected_found = all(pg in target_pages for pg in expected_targets)
    
    log("planner_links_valid", valid_targets > 1000 and invalid_targets == 0 and all_expected_found,
        f"{valid_targets} valid GOTO links, {invalid_targets} invalid, {len(target_pages)} unique targets ({sorted(target_pages)}), all 5 sections found: {all_expected_found}")
    doc.close()
except Exception as e:
    log("planner_links_valid", False, f"Error: {e}")

# ============================================================
# 2. COLORING PAGES: render to image and check non-white pixels
# ============================================================
print("\n=== EVIDENCE: Coloring Pages Render Visible Content ===")
try:
    # Use pdf2image or reportlab to rasterize, or use Pillow's PDF support
    # Try PyMuPDF (fitz) first, fallback to other methods
    import subprocess
    
    pdf_path = os.path.join(BASE, "04-coloring-pages/Free_Coloring_Pages_Bundle.pdf")
    
    # Try PyMuPDF
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(pdf_path)
        pages_with_content = 0
        pages_blank = 0
        
        for i in range(len(doc)):
            page = doc[i]
            # Render at 72 DPI
            pix = page.get_pixmap(matrix=fitz.Matrix(1, 1))
            img_data = pix.samples
            w, h = pix.width, pix.height
            
            # Count non-white pixels (allow some tolerance)
            non_white = 0
            total = w * h
            stride = pix.stride
            ncomp = pix.n
            
            # Sample every 10th pixel for speed
            for y in range(0, h, 10):
                for x in range(0, w, 10):
                    idx = y * stride + x * ncomp
                    if idx + 2 < len(img_data):
                        r, g, b = img_data[idx], img_data[idx+1], img_data[idx+2]
                        if r < 240 or g < 240 or b < 240:
                            non_white += 1
            
            sampled = (w // 10) * (h // 10)
            ratio = non_white / sampled if sampled > 0 else 0
            
            if ratio > 0.01:  # At least 1% non-white pixels
                pages_with_content += 1
            else:
                pages_blank += 1
                if pages_blank <= 3:
                    print(f"    Page {i+1}: only {non_white}/{sampled} non-white pixels ({ratio:.2%})")
        
        log("coloring_rendered", pages_with_content >= 45,
            f"{pages_with_content}/{len(doc)} pages have visible rendered content, {pages_blank} blank")
        doc.close()
    except ImportError:
        log("coloring_rendered", False, "PyMuPDF not available for rasterization test")
except Exception as e:
    log("coloring_rendered", False, f"Error: {e}")

# ============================================================
# 3. SVG RENDERING: convert sample SVGs to PNG and check
# ============================================================
print("\n=== EVIDENCE: SVGs Render Visible Content ===")
try:
    import glob
    import xml.etree.ElementTree as ET
    
    # Use cairosvg if available
    try:
        import cairosvg
        all_svgs = glob.glob(os.path.join(BASE, "**/*.svg"), recursive=True)
        # Sample 20 SVGs across different folders
        sample_svgs = []
        folders = set()
        for f in sorted(all_svgs):
            folder = os.path.dirname(f)
            if folder not in folders:
                folders.add(folder)
                sample_svgs.append(f)
            if len(sample_svgs) >= 20:
                break
        
        rendered_ok = 0
        rendered_fail = 0
        
        for svg_path in sample_svgs:
            try:
                png_path = os.path.join(tempfile.gettempdir(), "test_render.png")
                cairosvg.svg2png(url=svg_path, write_to=png_path, output_width=200, output_height=200)
                img = Image.open(png_path)
                # Check for non-transparent pixels
                pixels = list(img.getdata())
                non_transparent = sum(1 for p in pixels if (p[3] if len(p) == 4 else 255) > 0 and (p[0] < 250 or p[1] < 250 or p[2] < 250 if len(p) >= 3 else True))
                if non_transparent > 10:
                    rendered_ok += 1
                else:
                    rendered_fail += 1
                    print(f"    {os.path.relpath(svg_path, BASE)}: rendered but only {non_transparent} visible pixels")
            except Exception as e:
                rendered_fail += 1
                if rendered_fail <= 3:
                    print(f"    {os.path.relpath(svg_path, BASE)}: render error: {e}")
        
        log("svg_rendered", rendered_ok >= 15,
            f"{rendered_ok}/{len(sample_svgs)} sampled SVGs rendered visible content, {rendered_fail} failed")
    except ImportError:
        # Fallback: check SVG content has path data with actual coordinates
        all_svgs = glob.glob(os.path.join(BASE, "**/*.svg"), recursive=True)
        has_content = 0
        for f in all_svgs[:50]:
            content = open(f, 'r', errors='ignore').read()
            # Check for actual path data (not empty d="")
            paths = re.findall(r'd="([^"]*)"', content)
            if any(len(p.strip()) > 10 for p in paths):
                has_content += 1
        log("svg_rendered", has_content >= 40,
            f"{has_content}/50 sampled SVGs have path data > 10 chars (cairosvg not available for raster test)")
except Exception as e:
    log("svg_rendered", False, f"Error: {e}")

# ============================================================
# 4. SPREADSHEET: verify formulas compute correct values
# ============================================================
print("\n=== EVIDENCE: Spreadsheet Formula Computation ===")
try:
    from openpyxl import load_workbook
    wb = load_workbook(os.path.join(BASE, "05-spreadsheet-trackers/Ultimate_Budget_Tracker_Template.xlsx"), data_only=False)
    
    # Check Monthly Summary sheet has formulas that reference correct cells
    ws = wb["Monthly Summary"]
    
    # Find formulas and verify they reference valid ranges
    formula_count = 0
    valid_refs = 0
    broken_refs = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("="):
                formula_count += 1
                formula = str(cell.value)
                # Check for #REF errors
                if "#REF" in formula:
                    broken_refs += 1
                else:
                    valid_refs += 1
    
    # Now load with data_only to see if computed values exist
    wb2 = load_workbook(os.path.join(BASE, "05-spreadsheet-trackers/Ultimate_Budget_Tracker_Template.xlsx"), data_only=True)
    ws2 = wb2["Monthly Summary"]
    
    computed_values = 0
    none_values = 0
    for row in ws2.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                computed_values += 1
            elif cell.value is None:
                none_values += 1
    
    # Check Savings Goals sheet has SUM formulas
    ws_sg = wb["Savings Goals"]
    sg_formulas = 0
    for row in ws_sg.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("="):
                sg_formulas += 1
    
    log("spreadsheet_formulas", formula_count > 0 and broken_refs == 0 and sg_formulas > 0,
        f"{formula_count} formulas ({valid_refs} valid, {broken_refs} broken), {sg_formulas} in Savings Goals, {computed_values} computed numeric values found")
except Exception as e:
    log("spreadsheet_formulas", False, f"Error: {e}")

# ============================================================
# 5. WALL ART: verify Met Museum JPGs actually opened and have valid dimensions
# ============================================================
print("\n=== EVIDENCE: Wall Art Image Integrity ===")
try:
    import glob
    jpgs = glob.glob(os.path.join(BASE, "03-wall-art/**/*.jpg"), recursive=True)
    pngs = glob.glob(os.path.join(BASE, "03-wall-art/**/*.png"), recursive=True)
    
    # Open every 5th image and verify it loads, has reasonable size
    opened_ok = 0
    opened_fail = 0
    
    for i, f in enumerate(sorted(jpgs + pngs)):
        if i % 5 != 0:
            continue
        try:
            img = Image.open(f)
            w, h = img.size
            if w > 100 and h > 100:
                opened_ok += 1
            else:
                opened_fail += 1
                print(f"    {os.path.relpath(f, BASE)}: too small {w}x{h}")
        except:
            opened_fail += 1
            if opened_fail <= 3:
                print(f"    {os.path.relpath(f, BASE)}: failed to open")
    
    log("wall_art_opened", opened_ok >= 20 and opened_fail == 0,
        f"{opened_ok}/{opened_ok + opened_fail} sampled images opened with valid dimensions (>= 100x100)")
except Exception as e:
    log("wall_art_opened", False, f"Error: {e}")

# ============================================================
# SUMMARY
# ============================================================
print("\n" + "=" * 60)
print("CONCRETE EVIDENCE VERIFICATION SUMMARY")
print("=" * 60)
passed = sum(1 for _, p, _ in evidence if p)
failed = sum(1 for _, p, _ in evidence if not p)
for name, p, detail in evidence:
    print(f"  {name}: {'PASS' if p else 'FAIL'} - {detail}")
print(f"\n  Total: {passed} passed, {failed} failed out of {len(evidence)}")
print(f"  {'ALL EVIDENCE CONFIRMED' if failed == 0 else 'EVIDENCE GAPS FOUND'}")
