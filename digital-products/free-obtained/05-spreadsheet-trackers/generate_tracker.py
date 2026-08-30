"""
Generate a complete budget tracker spreadsheet (Excel/Google Sheets compatible).
Replaces Etsy listings selling "$7.99 ADHD Life Planner Budget Tracker" etc.

Output: .xlsx file with multiple sheets, formulas, conditional formatting, charts.
Requires: openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.dimensions import ColumnDimension
import os
from datetime import datetime

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Ultimate_Budget_Tracker_Template.xlsx")

# Colors
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E8B4B8", end_color="E8B4B8", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F5F2ED", end_color="F5F2ED", fill_type="solid")
INCOME_FILL = PatternFill(start_color="B8D8BA", end_color="B8D8BA", fill_type="solid")
EXPENSE_FILL = PatternFill(start_color="E8B4B8", end_color="E8B4B8", fill_type="solid")
SAVINGS_FILL = PatternFill(start_color="A8D5BA", end_color="A8D5BA", fill_type="solid")

HEADER_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="2C3E50")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="666666")
LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="2C3E50")
INPUT_FONT = Font(name="Calibri", size=10, color="333333")
FORMULA_FONT = Font(name="Calibri", size=10, color="666666", italic=True)

BOX_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

wb = openpyxl.Workbook()

# ============================
# Sheet 1: Dashboard
# ============================
ws = wb.active
ws.title = "Dashboard"

ws["B2"] = "💰 Ultimate Budget Tracker"
ws["B2"].font = TITLE_FONT
ws["B3"] = f"Created: {datetime.now().strftime('%Y-%m-%d')} | Free open-source template"
ws["B3"].font = SUBTITLE_FONT

# Quick stats
ws["B5"] = "Overview"
ws["B5"].font = LABEL_FONT
ws["B5"].fill = ACCENT_FILL
ws.merge_cells("B5:F5")

stats = [
    ("Total Income", "=SUM(Income!C:C)/2", INCOME_FILL, "$#,##0.00"),
    ("Total Expenses", "=SUM(Expenses!E:E)", EXPENSE_FILL, "$#,##0.00"),
    ("Net Balance", "=B6-B7", LIGHT_FILL, "$#,##0.00"),
    ("Savings Rate", "=IF(B6>0,B8/B6,0)", SAVINGS_FILL, "0.0%"),
    ("Monthly Budget", "=Settings!B2", LIGHT_FILL, "$#,##0.00"),
    ("Budget Remaining", "=B10-B7", SAVINGS_FILL, "$#,##0.00"),
]

for i, (label, formula, fill, num_fmt) in enumerate(stats):
    row = 6 + i
    ws[f"B{row}"] = label
    ws[f"B{row}"].font = LABEL_FONT
    ws[f"B{row}"].fill = fill
    ws[f"B{row}"].border = BOX_BORDER
    ws[f"C{row}"] = formula
    ws[f"C{row}"].font = Font(name="Calibri", bold=True, size=12)
    ws[f"C{row}"].number_format = num_fmt
    ws[f"C{row}"].border = BOX_BORDER
    ws.merge_cells(f"C{row}:F{row}")

# Category breakdown
ws["B14"] = "Expense by Category"
ws["B14"].font = LABEL_FONT
ws["B14"].fill = ACCENT_FILL
ws.merge_cells("B14:F14")

categories = ["Housing", "Food", "Transport", "Utilities", "Entertainment", "Health", "Shopping", "Other"]
ws["B15"] = "Category"
ws["C15"] = "Amount"
ws["D15"] = "% of Total"
for col in ["B", "C", "D"]:
    ws[f"{col}15"].font = LABEL_FONT
    ws[f"{col}15"].fill = LIGHT_FILL
    ws[f"{col}15"].border = BOX_BORDER

for i, cat in enumerate(categories):
    row = 16 + i
    ws[f"B{row}"] = cat
    ws[f"B{row}"].border = BOX_BORDER
    ws[f"C{row}"] = f'=SUMIF(Expenses!C:C,"{cat}",Expenses!E:E)'
    ws[f"C{row}"].number_format = "$#,##0.00"
    ws[f"C{row}"].border = BOX_BORDER
    ws[f"D{row}"] = f'=IF($C$23>0,C{row}/$C$23,0)'
    ws[f"D{row}"].number_format = "0.0%"
    ws[f"D{row}"].border = BOX_BORDER

ws["B24"] = "TOTAL"
ws["B24"].font = LABEL_FONT
ws["C24"] = "=SUM(C16:C23)"
ws["C24"].font = LABEL_FONT
ws["C24"].number_format = "$#,##0.00"
ws["D24"] = "=SUM(D16:D23)"
ws["D24"].number_format = "0.0%"
ws["B24"].border = BOX_BORDER
ws["C24"].border = BOX_BORDER
ws["D24"].border = BOX_BORDER

# Fix the expense total reference
ws["C23"] = ws["C23"].value  # Keep
ws["D23"] = ws["D23"].value
ws["C24"] = "=SUM(C16:C23)"
ws["D24"] = "=SUM(D16:D23)"

# Column widths
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 15

# ============================
# Sheet 2: Income
# ============================
ws2 = wb.create_sheet("Income")
ws2["A1"] = "💰 Income Tracker"
ws2["A1"].font = TITLE_FONT

ws2["A3"] = "Date"
ws2["B3"] = "Source"
ws2["C3"] = "Amount"
ws2["D3"] = "Notes"
for col in ["A", "B", "C", "D"]:
    ws2[f"{col}3"].font = HEADER_FONT
    ws2[f"{col}3"].fill = HEADER_FILL
    ws2[f"{col}3"].border = BOX_BORDER

# Sample rows
sample_income = [
    ("2026-01-01", "Salary", 3500, "Monthly paycheck"),
    ("2026-01-15", "Side hustle", 450, "Freelance design"),
    ("2026-02-01", "Salary", 3500, "Monthly paycheck"),
    ("2026-02-20", "Investment", 120, "Dividends"),
]
for i, (date, source, amount, notes) in enumerate(sample_income):
    row = 4 + i
    ws2[f"A{row}"] = date
    ws2[f"B{row}"] = source
    ws2[f"C{row}"] = amount
    ws2[f"C{row}"].number_format = "$#,##0.00"
    ws2[f"D{row}"] = notes
    for col in ["A", "B", "C", "D"]:
        ws2[f"{col}{row}"].border = BOX_BORDER
        ws2[f"{col}{row}"].font = INPUT_FONT

ws2.column_dimensions["A"].width = 15
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 15
ws2.column_dimensions["D"].width = 40

# ============================
# Sheet 3: Expenses
# ============================
ws3 = wb.create_sheet("Expenses")
ws3["A1"] = "💸 Expense Tracker"
ws3["A1"].font = TITLE_FONT

headers3 = ["Date", "Description", "Category", "Payment Method", "Amount", "Notes"]
for i, h in enumerate(headers3):
    col = get_column_letter(i + 1)
    ws3[f"{col}3"] = h
    ws3[f"{col}3"].font = HEADER_FONT
    ws3[f"{col}3"].fill = HEADER_FILL
    ws3[f"{col}3"].border = BOX_BORDER

# Data validation for category
from openpyxl.worksheet.datavalidation import DataValidation
cat_dv = DataValidation(type="list", formula1='"Housing,Food,Transport,Utilities,Entertainment,Health,Shopping,Other"', allow_blank=True)
ws3.add_data_validation(cat_dv)
cat_dv.add(f"C4:C1000")

sample_expenses = [
    ("2026-01-03", "Rent", "Housing", "Bank", 1200, "Monthly rent"),
    ("2026-01-05", "Grocery run", "Food", "Card", 180, "Weekly groceries"),
    ("2026-01-07", "Gas", "Transport", "Card", 45, "Fill up"),
    ("2026-01-10", "Electric bill", "Utilities", "Bank", 85, "Monthly"),
    ("2026-01-12", "Movie night", "Entertainment", "Cash", 30, "Date night"),
    ("2026-01-15", "Gym membership", "Health", "Card", 35, "Monthly"),
    ("2026-01-20", "New shoes", "Shopping", "Card", 75, "Running shoes"),
    ("2026-02-03", "Rent", "Housing", "Bank", 1200, "Monthly rent"),
    ("2026-02-05", "Grocery run", "Food", "Card", 195, "Weekly groceries"),
    ("2026-02-07", "Gas", "Transport", "Card", 42, "Fill up"),
]
for i, row_data in enumerate(sample_expenses):
    row = 4 + i
    for j, val in enumerate(row_data):
        col = get_column_letter(j + 1)
        ws3[f"{col}{row}"] = val
        ws3[f"{col}{row}"].border = BOX_BORDER
        ws3[f"{col}{row}"].font = INPUT_FONT
        if col == "E":
            ws3[f"{col}{row}"].number_format = "$#,##0.00"

ws3.column_dimensions["A"].width = 15
ws3.column_dimensions["B"].width = 25
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 15
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 30

# ============================
# Sheet 4: Savings Goals
# ============================
ws4 = wb.create_sheet("Savings Goals")
ws4["A1"] = "🎯 Savings Goals"
ws4["A1"].font = TITLE_FONT

headers4 = ["Goal", "Target Amount", "Saved So Far", "Remaining", "Progress %", "Target Date"]
for i, h in enumerate(headers4):
    col = get_column_letter(i + 1)
    ws4[f"{col}3"] = h
    ws4[f"{col}3"].font = HEADER_FONT
    ws4[f"{col}3"].fill = HEADER_FILL
    ws4[f"{col}3"].border = BOX_BORDER

sample_goals = [
    ("Emergency Fund", 10000, 3500, "2026-12-31"),
    ("Vacation", 3000, 850, "2026-06-01"),
    ("New Laptop", 2000, 1500, "2026-03-15"),
    ("House Down Payment", 50000, 8500, "2027-06-01"),
]
for i, (goal, target, saved, date) in enumerate(sample_goals):
    row = 4 + i
    ws4[f"A{row}"] = goal
    ws4[f"B{row}"] = target
    ws4[f"B{row}"].number_format = "$#,##0.00"
    ws4[f"C{row}"] = saved
    ws4[f"C{row}"].number_format = "$#,##0.00"
    ws4[f"D{row}"] = f"=B{row}-C{row}"
    ws4[f"D{row}"].number_format = "$#,##0.00"
    ws4[f"E{row}"] = f"=IF(B{row}>0,C{row}/B{row},0)"
    ws4[f"E{row}"].number_format = "0.0%"
    ws4[f"F{row}"] = date
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws4[f"{col}{row}"].border = BOX_BORDER
        ws4[f"{col}{row}"].font = INPUT_FONT

# Conditional formatting: progress bar
ws4.conditional_formatting.add("E4:E20",
    ColorScaleRule(start_type="num", start_value=0, start_color="E8B4B8",
                   end_type="num", end_value=1, end_color="B8D8BA"))

ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 15
ws4.column_dimensions["E"].width = 15
ws4.column_dimensions["F"].width = 15

# ============================
# Sheet 5: Settings
# ============================
ws5 = wb.create_sheet("Settings")
ws5["A1"] = "⚙️ Settings"
ws5["A1"].font = TITLE_FONT

ws5["A3"] = "Monthly Budget Limit"
ws5["A3"].font = LABEL_FONT
ws5["B3"] = 3000
ws5["B3"].number_format = "$#,##0.00"
ws5["B3"].fill = LIGHT_FILL
ws5["B3"].border = BOX_BORDER

ws5["A4"] = "Currency"
ws5["A4"].font = LABEL_FONT
ws5["B4"] = "USD"
ws5["B4"].border = BOX_BORDER

ws5["A5"] = "Start of Month"
ws5["A5"].font = LABEL_FONT
ws5["B5"] = 1
ws5["B5"].border = BOX_BORDER

ws5.column_dimensions["A"].width = 25
ws5.column_dimensions["B"].width = 20

# ============================
# Sheet 6: Monthly Summary
# ============================
ws6 = wb.create_sheet("Monthly Summary")
ws6["A1"] = "📊 Monthly Summary"
ws6["A1"].font = TITLE_FONT

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
ws6["A3"] = "Month"
ws6["B3"] = "Income"
ws6["C3"] = "Expenses"
ws6["D3"] = "Savings"
ws6["E3"] = "Savings %"
for col in ["A", "B", "C", "D", "E"]:
    ws6[f"{col}3"].font = HEADER_FONT
    ws6[f"{col}3"].fill = HEADER_FILL
    ws6[f"{col}3"].border = BOX_BORDER

for i, m in enumerate(months):
    row = 4 + i
    ws6[f"A{row}"] = m
    ws6[f"A{row}"].font = LABEL_FONT
    ws6[f"A{row}"].border = BOX_BORDER
    # Placeholder formulas that reference Income and Expenses sheets
    ws6[f"B{row}"] = f'=SUMIFS(Income!C:C,Income!A:A,"2026-{i+1:02d}*")'
    ws6[f"B{row}"].number_format = "$#,##0.00"
    ws6[f"B{row}"].border = BOX_BORDER
    ws6[f"C{row}"] = f'=SUMIFS(Expenses!E:E,Expenses!A:A,"2026-{i+1:02d}*")'
    ws6[f"C{row}"].number_format = "$#,##0.00"
    ws6[f"C{row}"].border = BOX_BORDER
    ws6[f"D{row}"] = f"=B{row}-C{row}"
    ws6[f"D{row}"].number_format = "$#,##0.00"
    ws6[f"D{row}"].border = BOX_BORDER
    ws6[f"E{row}"] = f'=IF(B{row}>0,D{row}/B{row},0)'
    ws6[f"E{row}"].number_format = "0.0%"
    ws6[f"E{row}"].border = BOX_BORDER

# Add bar chart: Income vs Expenses by Month
chart = BarChart()
chart.type = "col"
chart.title = "Income vs Expenses by Month"
chart.y_axis.title = "Amount ($)"
chart.x_axis.title = "Month"
data = Reference(ws6, min_col=2, min_row=3, max_col=4, max_row=15)
cats = Reference(ws6, min_col=1, min_row=4, max_row=15)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 20
chart.height = 10
ws6.add_chart(chart, "G3")

# Add line chart: Savings rate trend
line_chart = LineChart()
line_chart.title = "Savings Rate Trend"
line_chart.y_axis.title = "Savings Rate (%)"
line_chart.x_axis.title = "Month"
savings_data = Reference(ws6, min_col=5, min_row=3, max_row=15)
line_chart.add_data(savings_data, titles_from_data=True)
line_chart.set_categories(cats)
line_chart.width = 20
line_chart.height = 10
ws6.add_chart(line_chart, "G25")

# Add pie chart: Expense categories on Dashboard
pie_chart = PieChart()
pie_chart.title = "Expense Breakdown by Category"
# Reference expense categories and amounts from the Expenses sheet (ws3)
# Use the Dashboard's category breakdown (ws) for the pie chart
pie_cats = Reference(ws, min_col=2, min_row=16, max_row=23)
pie_data = Reference(ws, min_col=3, min_row=15, max_row=23)
pie_chart.add_data(pie_data, titles_from_data=True)
pie_chart.set_categories(pie_cats)
pie_chart.width = 15
pie_chart.height = 10
ws.add_chart(pie_chart, "E14")

ws6.column_dimensions["A"].width = 10
ws6.column_dimensions["B"].width = 15
ws6.column_dimensions["C"].width = 15
ws6.column_dimensions["D"].width = 15
ws6.column_dimensions["E"].width = 15

# Save
wb.save(OUTPUT_FILE)
print(f"Generated: {OUTPUT_FILE}")
print(f"Sheets: Dashboard, Income, Expenses, Savings Goals, Settings, Monthly Summary")
print(f"Features: Formulas, Charts, Conditional Formatting, Data Validation")
print(f"Compatible with: Excel, Google Sheets, LibreOffice Calc")
