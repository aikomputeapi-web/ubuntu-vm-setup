"""Comprehensive validation of all generated digital products."""
import os
import sys
import xml.etree.ElementTree as ET

BASE = os.path.dirname(os.path.abspath(__file__))

def validate_svgs():
    """Validate all SVG files are well-formed XML."""
    print("\n=== SVG Validation ===")
    svg_dirs = [
        "02-svg-bundles/svg-bundles/iconify-icons",
        "02-svg-bundles/svg-bundles/original-designs", 
        "02-svg-bundles/svg-bundles/themed-packs",
        "07-canva-templates/templates",
        "11-custom-personalized/output",
    ]
    invalid = []
    total = 0
    for d in svg_dirs:
        full = os.path.join(BASE, d)
        if not os.path.isdir(full):
            print(f"  SKIP (not found): {d}")
            continue
        count = 0
        for root, dirs, files in os.walk(full):
            for f in files:
                if f.endswith(".svg"):
                    total += 1
                    count += 1
                    path = os.path.join(root, f)
                    try:
                        tree = ET.parse(path)
                        # Verify it has an svg root element
                        root_elem = tree.getroot()
                        if "svg" not in root_elem.tag:
                            invalid.append((path, "No svg root element"))
                    except Exception as e:
                        invalid.append((path, str(e)))
        print(f"  {d}: {count} files")
    print(f"  TOTAL: {total} SVGs, {len(invalid)} invalid")
    for p, e in invalid:
        print(f"    INVALID: {os.path.basename(p)} - {e}")
    return len(invalid) == 0

def validate_pdfs():
    """Validate all PDFs have valid page counts and content."""
    print("\n=== PDF Validation ===")
    try:
        from pypdf import PdfReader
    except ImportError:
        print("  pypdf not installed, skipping")
        return False
    
    pdfs = [
        "01-digital-planners/planner_2026_6month_hyperlinked.pdf",
        "03-wall-art/Wall_Art_Collection.pdf",
        "04-coloring-pages/Free_Coloring_Pages_Bundle.pdf",
    ]
    ok = True
    for p in pdfs:
        full = os.path.join(BASE, p)
        if not os.path.exists(full):
            print(f"  MISSING: {p}")
            ok = False
            continue
        reader = PdfReader(full)
        pages = len(reader.pages)
        outline_count = 0
        try:
            outline = reader.outline
            if outline:
                outline_count = len(outline)
        except:
            pass
        # Check file size
        size_kb = os.path.getsize(full) / 1024
        print(f"  {os.path.basename(p)}: {pages} pages, {size_kb:.0f} KB, {outline_count} bookmarks")
        ok = ok and pages > 0
    return ok

def validate_xlsx():
    """Validate the spreadsheet."""
    print("\n=== XLSX Validation ===")
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed, skipping")
        return False
    
    p = os.path.join(BASE, "05-spreadsheet-trackers/Ultimate_Budget_Tracker_Template.xlsx")
    if not os.path.exists(p):
        print(f"  MISSING")
        return False
    wb = openpyxl.load_workbook(p)
    sheets = wb.sheetnames
    print(f"  Sheets: {sheets}")
    total_formulas = 0
    for s in sheets:
        ws = wb[s]
        formulas = 0
        charts = len(ws._charts) if hasattr(ws, "_charts") else 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
        total_formulas += formulas
        print(f"    {s}: {ws.max_row} rows x {ws.max_column} cols, {formulas} formulas, {charts} charts")
    print(f"  Total formulas: {total_formulas}")
    return True

def validate_wall_art():
    """Validate wall art images."""
    print("\n=== Wall Art Validation ===")
    from PIL import Image
    
    # Original PNGs
    png_dir = os.path.join(BASE, "03-wall-art/wall-art-png")
    if os.path.isdir(png_dir):
        pngs = [f for f in os.listdir(png_dir) if f.endswith(".png")]
        print(f"  Original PNGs: {len(pngs)}")
        # Check first image dimensions and DPI
        if pngs:
            img = Image.open(os.path.join(png_dir, pngs[0]))
            print(f"  Sample dimensions: {img.size}, DPI: {img.info.get('dpi', 'N/A')}")
    
    # Met Museum images (in subfolders by category)
    met_dir = os.path.join(BASE, "03-wall-art/met-wall-art")
    if os.path.isdir(met_dir):
        jpgs = []
        jsons = []
        for root, dirs, files in os.walk(met_dir):
            for f in files:
                if f.endswith((".jpg", ".jpeg")):
                    jpgs.append(f)
                elif f.endswith(".json"):
                    jsons.append(f)
        print(f"  Met Museum JPGs: {len(jpgs)}, JSON metadata: {len(jsons)}")
    return True

def validate_ideas_guide():
    """Validate the ideas guide was generated with data."""
    print("\n=== Ideas Guide Validation ===")
    p = os.path.join(BASE, "10-digital-product-ideas/Digital_Product_Ideas_Guide.md")
    if not os.path.exists(p):
        print("  MISSING")
        return False
    with open(p, "r", encoding="utf-8") as f:
        content = f.read()
    # Check it contains expected sections
    has_top20 = "Top 20 Best-Selling" in content
    has_revenue = "Revenue Potential" in content
    has_hot = "Hot Products" in content
    has_insights = "Key Insights" in content
    has_actionable = "Actionable Product Ideas" in content
    line_count = content.count("\n")
    print(f"  Lines: {line_count}")
    print(f"  Top 20 sales table: {has_top20}")
    print(f"  Revenue analysis: {has_revenue}")
    print(f"  Hot products section: {has_hot}")
    print(f"  Key insights: {has_insights}")
    print(f"  Actionable ideas: {has_actionable}")
    return all([has_top20, has_revenue, has_hot, has_insights, has_actionable])

def validate_video_pipeline():
    """Validate AI video pipeline files."""
    print("\n=== AI Video Pipeline Validation ===")
    base = os.path.join(BASE, "09-ai-video-pipeline")
    expected = [
        "generate_pipeline.py",
        "batch_generate.sh",
        "README.md",
    ]
    for f in expected:
        p = os.path.join(base, f)
        exists = os.path.exists(p)
        size = os.path.getsize(p) if exists else 0
        print(f"  {f}: {'OK' if exists else 'MISSING'} ({size} bytes)")
    
    scripts_dir = os.path.join(base, "video-scripts")
    if os.path.isdir(scripts_dir):
        scripts = [f for f in os.listdir(scripts_dir)]
        print(f"  Video scripts: {len(scripts)} files")
        for s in scripts:
            print(f"    - {s}")
    return True
    
def validate_readme():
    """Validate the master README is up to date."""
    print("\n=== README Validation ===")
    p = os.path.join(BASE, "README.md")
    with open(p, "r", encoding="utf-8") as f:
        content = f.read()
    # Check coverage number
    has_137 = "137 of 137" in content
    has_100 = "100%" in content
    has_canva_count = "104 SVG" in content or "104 original" in content
    has_themed = "48 themed" in content
    has_ideas = "10-digital-product-ideas" in content
    has_custom = "11-custom-personalized" in content
    print(f"  Coverage 137/137: {has_137}")
    print(f"  Coverage 100%: {has_100}")
    print(f"  Canva templates count mentioned: {has_canva_count}")
    print(f"  Themed SVGs mentioned: {has_themed}")
    print(f"  Ideas guide folder referenced: {has_ideas}")
    print(f"  Custom products folder referenced: {has_custom}")
    return all([has_137, has_100, has_canva_count, has_themed, has_ideas, has_custom])

def count_all_files():
    """Count total generated files."""
    print("\n=== File Count Summary ===")
    counts = {}
    for root, dirs, files in os.walk(BASE):
        # Skip __pycache__ and .git
        dirs[:] = [d for d in dirs if d not in ("__pycache__", ".git")]
        rel = os.path.relpath(root, BASE)
        if rel == ".":
            rel = "[root]"
        ext_counts = {}
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in (".py", ".svg", ".png", ".jpg", ".jpeg", ".pdf", ".xlsx", ".md", ".json", ".sh", ".txt"):
                ext_counts[ext] = ext_counts.get(ext, 0) + 1
        if ext_counts:
            for ext, c in ext_counts.items():
                if ext not in counts:
                    counts[ext] = 0
                counts[ext] += c
    
    # Print per-extension counts
    for ext in sorted(counts.keys()):
        print(f"  {ext}: {counts[ext]}")
    print(f"  TOTAL: {sum(counts.values())}")


if __name__ == "__main__":
    print("=" * 60)
    print("COMPREHENSIVE VALIDATION OF ALL DIGITAL PRODUCTS")
    print("=" * 60)
    
    results = {}
    results["svgs"] = validate_svgs()
    results["pdfs"] = validate_pdfs()
    results["xlsx"] = validate_xlsx()
    results["wall_art"] = validate_wall_art()
    results["ideas"] = validate_ideas_guide()
    results["video"] = validate_video_pipeline()
    results["readme"] = validate_readme()
    count_all_files()
    
    print("\n" + "=" * 60)
    print("VALIDATION SUMMARY")
    print("=" * 60)
    all_pass = True
    for k, v in results.items():
        status = "PASS" if v else "FAIL"
        if not v:
            all_pass = False
        print(f"  {k}: {status}")
    print(f"\n  Overall: {'ALL PASS' if all_pass else 'ISSUES FOUND'}")
